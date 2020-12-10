import openpyxl
import os

os.chdir(r'C:\Users\WayneJnr\Desktop')#Changing directory to worksheet.

workbook = openpyxl.load_workbook('example.xlsx') #Opening workbook.

sheet = workbook['Sheet1'] #Opening specific sheet.

print(sheet['a1'].value) #How to print out a cells value.

for i in range(1, 8): #For loop to print out all of the data in column 1 and 2
    print(i, sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)

#Writing to excel sheets.
wb = openpyxl.load_workbook('Example2.xlsx')
sheet=wb['Sheet']
sheet['A1'] = 56
sheet['A2'] = 57
sheet['A3'] = 58
print(sheet['A1'].value,sheet['A2'].value,sheet['A3'].value)
wb.create_sheet('SheetFour')#Creation of a new sheet.
print(wb.sheetnames) #Print Sheetnames
#wb.save('Example2.xlsx') Saving changes.
